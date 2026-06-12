"""
메가스터디 입시정보 크롤러
- 입시뉴스, 입시대담, 교육기관발표자료 수집
- Gemini API로 구조화 데이터 추출
- Neon PostgreSQL 저장
"""

import os
import re
import io
import time
import json
import zlib
import struct
import zipfile
import logging
import tempfile
from datetime import datetime, date
from typing import Optional
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup
import psycopg2
from psycopg2.extras import Json, execute_values
from dotenv import load_dotenv

load_dotenv(dotenv_path=os.path.join(os.path.dirname(__file__), '..', '.env.local'))

logging.basicConfig(level=logging.INFO, format='%(asctime)s %(levelname)s %(message)s')
log = logging.getLogger(__name__)

DATABASE_URL = os.environ['DATABASE_URL']
OPENROUTER_API_KEY = os.environ.get('OPENROUTER_API_KEY', '')
OPENROUTER_MODEL = 'openai/gpt-oss-120b:free'
HF_EMBED_URL = (
    'https://api-inference.huggingface.co/pipeline/feature-extraction/'
    'sentence-transformers/paraphrase-multilingual-mpnet-base-v2'
)

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'ko-KR,ko;q=0.9,en;q=0.8',
    'Referer': 'https://www.megastudy.net/',
}

ATTACH_EXTS = {'.pdf', '.hwp', '.hwpx', '.xlsx', '.xls', '.docx', '.doc'}

SOURCES = {
    'news': {
        'label': '입시뉴스',
        'list_url': 'https://www.megastudy.net/Entinfo/News/news_list_ax.asp?page={page}',
        'view_url': 'https://www.megastudy.net/Entinfo/News/news_view_ax.asp?idx={idx}',
        'onclick_pattern': r'fncNewsView\((\d+)\)',
        'base_url': 'https://www.megastudy.net/Entinfo/News',
    },
    'ipsi_news': {
        'label': '입시대담',
        'list_url': 'https://www.megastudy.net/Entinfo/ipsi_News/news_list_ax.asp?page={page}',
        'view_url': 'https://www.megastudy.net/Entinfo/ipsi_News/news_view_ax.asp?idx={idx}',
        'onclick_pattern': r'fncNewsView\((\d+)\)',
        'base_url': 'https://www.megastudy.net/Entinfo/ipsi_News',
    },
    'archive': {
        'label': '교육기관발표자료',
        'list_url': 'https://www.megastudy.net/entinfo/g_archive/list_ax.asp?pidx={page}',
        'view_url': 'https://www.megastudy.net/entinfo/g_archive/view_ax.asp?idx={idx}',
        'onclick_pattern': r'fncAchivView\((\d+)\)',
        'base_url': 'https://www.megastudy.net/entinfo/g_archive',
    },
}


# ── 첨부파일 처리 ─────────────────────────────────────────────

def find_attachments(soup: BeautifulSoup, page_url: str) -> list[tuple[str, str]]:
    """첨부파일 (url, filename) 목록 반환."""
    seen: set[str] = set()
    results: list[tuple[str, str]] = []

    for a in soup.find_all('a', href=True):
        href = a['href'].strip()
        if not href or href.startswith('#'):
            continue

        link_text = a.get_text(strip=True)
        href_path = urlparse(href).path.lower()
        ext = os.path.splitext(href_path)[1]
        text_ext = os.path.splitext(link_text.lower())[1] if link_text else ''

        actual_ext = ext if ext in ATTACH_EXTS else (text_ext if text_ext in ATTACH_EXTS else None)
        if actual_ext:
            full_url = urljoin(page_url, href)
            if full_url not in seen:
                seen.add(full_url)
                filename = link_text if text_ext in ATTACH_EXTS else (os.path.basename(href_path) or 'attachment')
                results.append((full_url, filename))

    # onclick 속성에 파일 경로가 있는 경우
    for tag in soup.find_all(onclick=True):
        onclick = tag['onclick']
        m = re.search(r'["\']([^"\']+\.(hwp|hwpx|pdf|xlsx|xls|docx))["\']', onclick, re.IGNORECASE)
        if m:
            full_url = urljoin(page_url, m.group(1))
            if full_url not in seen:
                seen.add(full_url)
                results.append((full_url, os.path.basename(m.group(1))))

    return results


def download_file(url: str) -> Optional[bytes]:
    try:
        resp = requests.get(url, headers=HEADERS, timeout=30)
        if resp.status_code == 200:
            return resp.content
        log.warning(f"파일 다운로드 HTTP {resp.status_code}: {url}")
    except Exception as e:
        log.warning(f"파일 다운로드 실패 ({url}): {e}")
    return None


def _extract_pdf(data: bytes) -> str:
    import pdfplumber
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        return '\n'.join(p.extract_text() or '' for p in pdf.pages)


def _hwp_parse_records(data: bytes) -> str:
    """HWP5 BodyText 레코드에서 텍스트 파싱."""
    chars: list[str] = []
    pos = 0
    while pos + 4 <= len(data):
        hdr = struct.unpack_from('<I', data, pos)[0]
        rec_type = hdr & 0x3FF
        size = (hdr >> 20) & 0xFFF
        pos += 4
        if size == 0xFFF:
            if pos + 4 > len(data):
                break
            size = struct.unpack_from('<I', data, pos)[0]
            pos += 4
        rec = data[pos:pos + size]
        pos += size
        if rec_type == 67:  # HWPTAG_PARA_TEXT
            for j in range(0, len(rec) - 1, 2):
                c = struct.unpack_from('<H', rec, j)[0]
                if c == 13:
                    chars.append('\n')
                elif c in (9, 32):
                    chars.append(' ')
                elif 0x20 <= c <= 0xFFFF:
                    try:
                        chars.append(chr(c))
                    except Exception:
                        pass
    return ''.join(chars)


def _extract_hwp(data: bytes) -> str:
    """HWP5 (OLE 기반) 텍스트 추출."""
    try:
        import olefile
        ole = olefile.OleFileIO(io.BytesIO(data))
        texts: list[str] = []
        for i in range(200):
            stream = f'BodyText/Section{i:04d}'
            if not ole.exists(stream):
                break
            raw = ole.openstream(stream).read()
            try:
                raw = zlib.decompress(raw, -15)
            except zlib.error:
                pass
            texts.append(_hwp_parse_records(raw))
        ole.close()
        return '\n'.join(t for t in texts if t)
    except Exception as e:
        log.warning(f"HWP 파싱 실패: {e}")
        return ''


def _extract_hwpx(data: bytes) -> str:
    """HWPX (ZIP 기반) 텍스트 추출."""
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            texts: list[str] = []
            for name in sorted(zf.namelist()):
                if 'section' in name.lower() and name.endswith('.xml'):
                    content = zf.read(name).decode('utf-8', errors='ignore')
                    text = re.sub(r'<[^>]+>', ' ', content)
                    text = re.sub(r'\s+', ' ', text).strip()
                    if text:
                        texts.append(text)
            return '\n'.join(texts)
    except Exception as e:
        log.warning(f"HWPX 파싱 실패: {e}")
        return ''


def _extract_excel(data: bytes) -> str:
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    rows: list[str] = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            row_str = '\t'.join(str(v) for v in row if v is not None)
            if row_str.strip():
                rows.append(row_str)
    wb.close()
    return '\n'.join(rows)


def _extract_docx(data: bytes) -> str:
    from docx import Document
    doc = Document(io.BytesIO(data))
    return '\n'.join(p.text for p in doc.paragraphs if p.text.strip())


def extract_attachment_text(data: bytes, filename: str) -> str:
    ext = os.path.splitext(filename.lower())[1]
    try:
        if ext == '.pdf':
            return _extract_pdf(data)
        elif ext == '.hwp':
            return _extract_hwp(data)
        elif ext == '.hwpx':
            return _extract_hwpx(data)
        elif ext == '.xlsx':
            return _extract_excel(data)
        elif ext == '.docx':
            return _extract_docx(data)
        elif ext == '.xls':
            try:
                import xlrd
                wb = xlrd.open_workbook(file_contents=data)
                rows = []
                for sheet in wb.sheets():
                    for rx in range(sheet.nrows):
                        row_str = '\t'.join(str(sheet.cell(rx, cx).value) for cx in range(sheet.ncols))
                        if row_str.strip():
                            rows.append(row_str)
                return '\n'.join(rows)
            except ImportError:
                return ''
    except Exception as e:
        log.warning(f"첨부파일 텍스트 추출 실패 ({filename}): {e}")
    return ''


def get_db():
    conn = psycopg2.connect(DATABASE_URL)
    conn.autocommit = False
    return conn


def fetch_html(url: str, retries: int = 3) -> Optional[BeautifulSoup]:
    for attempt in range(retries):
        try:
            resp = requests.get(url, headers=HEADERS, timeout=20)
            # 사이트는 EUC-KR / CP949 인코딩 사용
            resp.encoding = resp.apparent_encoding or 'cp949'
            if resp.status_code == 200:
                return BeautifulSoup(resp.text, 'lxml')
            log.warning(f"HTTP {resp.status_code}: {url}")
        except Exception as e:
            log.warning(f"Attempt {attempt+1} failed for {url}: {e}")
            time.sleep(2 ** attempt)
    return None


def get_article_ids_from_page(source_key: str, page: int) -> list[int]:
    src = SOURCES[source_key]
    url = src['list_url'].format(page=page)
    soup = fetch_html(url)
    if not soup:
        return []

    pattern = src['onclick_pattern']
    ids = []
    for tag in soup.find_all(onclick=True):
        m = re.search(pattern, tag['onclick'])
        if m:
            ids.append(int(m.group(1)))
    return ids


def get_all_article_ids(source_key: str, max_pages: int = 50) -> list[int]:
    all_ids = []
    for page in range(1, max_pages + 1):
        ids = get_article_ids_from_page(source_key, page)
        if not ids:
            log.info(f"[{source_key}] 페이지 {page}에서 종료 (총 {len(all_ids)}개 발견)")
            break
        log.info(f"[{source_key}] 페이지 {page}: {len(ids)}개 발견")
        all_ids.extend(ids)
        time.sleep(1)
    return list(set(all_ids))


def parse_article(source_key: str, idx: int) -> Optional[dict]:
    src = SOURCES[source_key]
    url = src['view_url'].format(idx=idx)
    soup = fetch_html(url)
    if not soup:
        return None

    # 제목
    title_el = soup.find('h2') or soup.find('h1')
    title = title_el.get_text(strip=True) if title_el else ''

    # 날짜 (등록일 : 2026-05-27 패턴)
    full_text = soup.get_text()
    date_match = re.search(r'등록일\s*:\s*(\d{4}-\d{2}-\d{2})', full_text)
    published_at = None
    if date_match:
        try:
            published_at = datetime.strptime(date_match.group(1), '%Y-%m-%d').date()
        except ValueError:
            pass

    # 본문 (.viewContents)
    content_el = soup.find(class_='viewContents')
    if not content_el:
        content_el = soup.find(class_='view_con') or soup.find(id='articleContent')

    raw_content = content_el.get_text(separator='\n', strip=True) if content_el else ''

    if not title and not raw_content:
        return None

    # 첨부파일 텍스트 추출
    attachments = find_attachments(soup, url)
    if attachments:
        log.info(f"  첨부파일 {len(attachments)}개 발견")
    for att_url, att_name in attachments[:5]:  # 최대 5개
        file_data = download_file(att_url)
        if not file_data:
            continue
        att_text = extract_attachment_text(file_data, att_name)
        if att_text.strip():
            raw_content += f'\n\n[첨부: {att_name}]\n{att_text[:6000]}'
            log.info(f"  → 첨부파일 텍스트 {len(att_text)}자 추출: {att_name}")
        time.sleep(0.5)

    return {
        'source': source_key,
        'source_idx': idx,
        'title': title,
        'url': url,
        'published_at': published_at,
        'raw_content': raw_content,
    }


def save_article(conn, article: dict) -> Optional[int]:
    with conn.cursor() as cur:
        cur.execute("""
            INSERT INTO articles (source, source_idx, title, url, published_at, raw_content)
            VALUES (%s, %s, %s, %s, %s, %s)
            ON CONFLICT (source, source_idx) DO UPDATE SET
                title = EXCLUDED.title,
                url = EXCLUDED.url,
                raw_content = EXCLUDED.raw_content
            RETURNING id
        """, (
            article['source'],
            article['source_idx'],
            article['title'],
            article['url'],
            article['published_at'],
            article['raw_content'],
        ))
        row = cur.fetchone()
        conn.commit()
        return row[0] if row else None


# ── AI 분석 (OpenRouter) ───────────────────────────────────────

def setup_ai():
    if not OPENROUTER_API_KEY:
        log.warning("OPENROUTER_API_KEY 없음 - AI 분석 건너뜀")
        return None
    log.info(f"OpenRouter 준비 완료 (모델: {OPENROUTER_MODEL})")
    return True  # 상태 플래그


ANALYSIS_PROMPT = """아래는 한국 대학입시 관련 기사/자료입니다. 다음 항목을 JSON으로 추출해주세요.

규칙:
- 없는 정보는 null 또는 빈 배열로
- universities: 언급된 대학교 이름 목록 (예: ["서울대", "연세대"])
- year: 해당 학년도 숫자 (예: 2027)
- admission_types: 전형 종류 목록 (예: ["수시", "정시", "논술", "학종", "교과"])
- cutoff_scores: 등급컷/점수 데이터 {{대학: {{전형: 점수}}}} 형태
- competition_rates: 경쟁률 데이터 {{대학: {{전형: 경쟁률}}}} 형태
- key_changes: 전년 대비 주요 변경사항 한 문단 (없으면 null)
- summary: 3줄 이내 핵심 요약

반드시 유효한 JSON만 출력하세요. 마크다운 코드블록 없이 순수 JSON.

기사:
---
{content}
---"""

ARCHIVE_PROMPT = """아래는 한국 교육기관이 발표한 공식 자료입니다(수능, 모의고사, 학력평가 등). 다음 항목을 JSON으로 추출해주세요.

규칙:
- 없는 정보는 null 또는 빈 배열로
- year: 시험 시행 연도 숫자 (예: 2024)
- exam_type: 시험 종류 문자열 (예: "수능", "전국연합학력평가", "모의고사", "대학별고사", "기타")
- applicant_count: 총 응시 인원 숫자 (없으면 null)
- universities: 언급된 대학교 목록
- admission_types: 관련 전형 목록
- summary: 3줄 이내 핵심 요약
- key_changes: 전년 대비 주요 변경사항 (없으면 null)
- exam_stats: 시험 세부 데이터 — 다음 형태로 최대한 추출:
  {{
    "subjects": {{
      "국어": {{"응시": 숫자, "평균": 숫자, "표준편차": 숫자, "등급컷": {{1: 점수, 2: 점수, ...}}}},
      "수학": {{...}},
      "영어": {{...}}
    }},
    "grade_distribution": {{"1등급": 숫자, "2등급": 숫자, ...}},
    "notes": "기타 주요 통계"
  }}

반드시 유효한 JSON만 출력하세요. 마크다운 코드블록 없이 순수 JSON.

자료:
---
{content}
---"""


def analyze_with_ai(enabled, article: dict) -> Optional[dict]:
    if not enabled or not OPENROUTER_API_KEY:
        return None

    content = f"제목: {article['title']}\n\n{article['raw_content']}"
    content = content[:8000]

    try:
        resp = requests.post(
            'https://openrouter.ai/api/v1/chat/completions',
            headers={
                'Authorization': f'Bearer {OPENROUTER_API_KEY}',
                'Content-Type': 'application/json',
            },
            json={
                'model': OPENROUTER_MODEL,
                'messages': [{'role': 'user', 'content': ANALYSIS_PROMPT.format(content=content)}],
                'temperature': 0.1,
            },
            timeout=60,
        )
        resp.raise_for_status()
        text = resp.json()['choices'][0]['message']['content'].strip()
        text = re.sub(r'^```(?:json)?\n?', '', text)
        text = re.sub(r'\n?```$', '', text)
        return json.loads(text)
    except Exception as e:
        log.warning(f"AI 분석 실패: {e}")
        return None


def save_structured(conn, article_id: int, data: dict):
    with conn.cursor() as cur:
        cur.execute("""
            INSERT INTO structured_data
              (article_id, universities, year, admission_types,
               cutoff_scores, competition_rates, key_changes, summary,
               exam_type, applicant_count, exam_stats)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT DO NOTHING
        """, (
            article_id,
            data.get('universities') or [],
            data.get('year'),
            data.get('admission_types') or [],
            Json(data.get('cutoff_scores') or {}),
            Json(data.get('competition_rates') or {}),
            data.get('key_changes'),
            data.get('summary'),
            data.get('exam_type'),
            data.get('applicant_count'),
            Json(data.get('exam_stats') or {}),
        ))
        conn.commit()


def save_structured_archive(conn, article_id: int, data: dict):
    """archive 전용 — 기존 레코드 UPDATE 또는 INSERT."""
    with conn.cursor() as cur:
        cur.execute("""
            INSERT INTO structured_data
              (article_id, universities, year, admission_types,
               cutoff_scores, competition_rates, key_changes, summary,
               exam_type, applicant_count, exam_stats)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ON CONFLICT (article_id) DO UPDATE SET
               universities = EXCLUDED.universities,
               year = EXCLUDED.year,
               admission_types = EXCLUDED.admission_types,
               key_changes = EXCLUDED.key_changes,
               summary = EXCLUDED.summary,
               exam_type = EXCLUDED.exam_type,
               applicant_count = EXCLUDED.applicant_count,
               exam_stats = EXCLUDED.exam_stats
        """, (
            article_id,
            data.get('universities') or [],
            data.get('year'),
            data.get('admission_types') or [],
            Json(data.get('cutoff_scores') or {}),
            Json(data.get('competition_rates') or {}),
            data.get('key_changes'),
            data.get('summary'),
            data.get('exam_type'),
            data.get('applicant_count'),
            Json(data.get('exam_stats') or {}),
        ))
        conn.commit()


# ── 임베딩 (HuggingFace) ─────────────────────────────────────

def get_embedding(text: str) -> Optional[list[float]]:
    for attempt in range(3):
        try:
            resp = requests.post(HF_EMBED_URL, json={'inputs': text}, timeout=30)
            if resp.status_code == 503:
                wait = min(resp.json().get('estimated_time', 20), 30)
                log.info(f"HuggingFace 모델 로딩 중... {wait:.0f}초 대기")
                time.sleep(wait)
                continue
            resp.raise_for_status()
            result = resp.json()
            # [[float, ...]] 또는 [float, ...]
            if isinstance(result, list) and isinstance(result[0], list):
                return result[0]
            return result
        except Exception as e:
            log.warning(f"임베딩 실패 (시도 {attempt+1}): {e}")
            time.sleep(5)
    return None


def chunk_text(text: str, chunk_size: int = 500, overlap: int = 50) -> list[str]:
    words = text.split()
    chunks = []
    i = 0
    while i < len(words):
        chunk = ' '.join(words[i:i + chunk_size])
        chunks.append(chunk)
        i += chunk_size - overlap
    return chunks


def save_embeddings(conn, article_id: int, content: str):
    chunks = chunk_text(content)
    rows = []
    for chunk in chunks[:10]:  # 최대 10 청크
        emb = get_embedding(chunk)
        if emb:
            rows.append((article_id, chunk, emb))
        time.sleep(0.1)

    if rows:
        with conn.cursor() as cur:
            execute_values(cur, """
                INSERT INTO search_embeddings (article_id, chunk_text, embedding)
                VALUES %s
                ON CONFLICT DO NOTHING
            """, rows, template='(%s, %s, %s::vector)')
        conn.commit()


# ── 메인 파이프라인 ────────────────────────────────────────────

def get_existing_ids(conn) -> set[tuple]:
    with conn.cursor() as cur:
        cur.execute("SELECT source, source_idx FROM articles WHERE source_idx IS NOT NULL")
        return {(row[0], row[1]) for row in cur.fetchall()}


def run(source_keys: list[str] = None, max_pages: int = 50, analyze: bool = True):
    if source_keys is None:
        source_keys = list(SOURCES.keys())

    conn = get_db()
    model = setup_ai() if analyze else None
    existing_ids = get_existing_ids(conn)

    for source_key in source_keys:
        label = SOURCES[source_key]['label']
        log.info(f"\n{'='*50}\n[{label}] 크롤링 시작\n{'='*50}")

        all_ids = get_all_article_ids(source_key, max_pages=max_pages)
        log.info(f"[{label}] 총 {len(all_ids)}개 ID 발견")

        new_count = 0
        for idx in sorted(all_ids, reverse=True):
            if (source_key, idx) in existing_ids:
                continue

            url = SOURCES[source_key]['view_url'].format(idx=idx)
            article = parse_article(source_key, idx)
            if not article:
                log.warning(f"파싱 실패: {url}")
                time.sleep(0.5)
                continue

            article_id = save_article(conn, article)
            if not article_id:
                continue

            log.info(f"[{label}] 저장: {article['title'][:50]}")
            new_count += 1
            existing_ids.add((source_key, idx))

            if model and article.get('raw_content'):
                structured = analyze_with_ai(model, article)
                if structured:
                    save_structured(conn, article_id, structured)
                    log.info(f"  → 구조화 데이터 저장")

                save_embeddings(conn, article_id, article['raw_content'])
                time.sleep(1)  # API rate limit
            else:
                time.sleep(0.3)

        log.info(f"[{label}] 신규 저장: {new_count}개")

    conn.close()
    log.info("크롤링 완료")


def fill_missing_analysis():
    """structured_data/임베딩이 없는 기존 기사를 소급 분석."""
    model = setup_ai()
    if not model:
        log.error("OPENROUTER_API_KEY 필요")
        return

    conn = get_db()
    with conn.cursor() as cur:
        cur.execute("""
            SELECT a.id, a.title, a.raw_content
            FROM articles a
            LEFT JOIN structured_data s ON s.article_id = a.id
            WHERE s.id IS NULL AND a.raw_content IS NOT NULL AND a.raw_content != ''
            ORDER BY a.id DESC
        """)
        articles = cur.fetchall()

    log.info(f"소급 처리 대상: {len(articles)}개 기사")
    for i, (article_id, title, raw_content) in enumerate(articles, 1):
        log.info(f"[{i}/{len(articles)}] {title[:50]}")

        structured = analyze_with_ai(model, {'title': title, 'raw_content': raw_content})
        if structured:
            save_structured(conn, article_id, structured)
            log.info(f"  → 구조화 데이터 저장")

        time.sleep(0.5)

    conn.close()
    log.info("소급 처리 완료")


def refill_archive_analysis():
    """교육기관발표자료를 전용 프롬프트로 재분석 (exam_type, applicant_count, exam_stats 추출)."""
    model = setup_ai()
    if not model:
        log.error("OPENROUTER_API_KEY 필요")
        return

    conn = get_db()
    with conn.cursor() as cur:
        cur.execute("""
            SELECT a.id, a.title, a.raw_content
            FROM articles a
            WHERE a.source = 'archive' AND a.raw_content IS NOT NULL AND a.raw_content != ''
            ORDER BY a.id DESC
        """)
        articles = cur.fetchall()

    log.info(f"교육기관발표자료 재분석 대상: {len(articles)}개")
    for i, (article_id, title, raw_content) in enumerate(articles, 1):
        log.info(f"[{i}/{len(articles)}] {title[:50]}")
        content = f"제목: {title}\n\n{raw_content}"[:8000]
        try:
            resp = requests.post(
                'https://openrouter.ai/api/v1/chat/completions',
                headers={'Authorization': f'Bearer {OPENROUTER_API_KEY}', 'Content-Type': 'application/json'},
                json={
                    'model': OPENROUTER_MODEL,
                    'messages': [{'role': 'user', 'content': ARCHIVE_PROMPT.format(content=content)}],
                    'temperature': 0.1,
                },
                timeout=60,
            )
            resp.raise_for_status()
            text = resp.json()['choices'][0]['message']['content'].strip()
            text = re.sub(r'^```(?:json)?\n?', '', text)
            text = re.sub(r'\n?```$', '', text)
            data = json.loads(text)
            save_structured_archive(conn, article_id, data)
            log.info(f"  → 저장 완료 (exam_type={data.get('exam_type')}, applicant={data.get('applicant_count')})")
        except Exception as e:
            log.warning(f"  재분석 실패: {e}")
        time.sleep(0.5)

    conn.close()
    log.info("교육기관발표자료 재분석 완료")


if __name__ == '__main__':
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument('--sources', nargs='+', choices=list(SOURCES.keys()),
                        help='수집할 소스 (기본: 전체)')
    parser.add_argument('--max-pages', type=int, default=50)
    parser.add_argument('--no-analyze', action='store_true', help='AI 분석 건너뜀')
    parser.add_argument('--fill-missing', action='store_true',
                        help='기존 기사 중 분석 누락된 것 소급 처리')
    parser.add_argument('--refill-archive', action='store_true',
                        help='교육기관발표자료 전용 프롬프트로 재분석')
    args = parser.parse_args()

    if args.fill_missing:
        fill_missing_analysis()
    elif args.refill_archive:
        refill_archive_analysis()
    else:
        run(
            source_keys=args.sources,
            max_pages=args.max_pages,
            analyze=not args.no_analyze,
        )
